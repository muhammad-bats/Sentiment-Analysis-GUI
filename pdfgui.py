import tkinter as tk
import PyPDF2 as PDF
import numpy as np
import openpyxl
import os
import os.path
import spacy
from tkinter import messagebox, filedialog
from nltk.tokenize import sent_tokenize 
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from textblob import TextBlob
from collections import Counter
from flair.models import TextClassifier
from flair.data import Sentence
from spacytextblob.spacytextblob import SpacyTextBlob


def select_model(option):
    global selected_model
    selected_model = option
    print(selected_model)

def show_error(level):
    if level == 1:
        messagebox.showerror("Error", "No folder selected")
    elif level == 2:
        messagebox.showerror("Error", "No file selected")
    else:
        messagebox.showerror("Error", "An error occurred!")

def get_colnum(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    if sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        max_col = sheet.max_column
        new_col = max_col + 2
    else:
        new_col = 3
    return new_col

def folder_directory():
    global folder_path
    folder_path = filedialog.askdirectory()
    print(folder_path)

def label_frequency(list):
    counts = Counter(list)
    max_count = max(counts.values())
    modes = [item for item, count in counts.items() if count == max_count]

    return modes[0] if len(modes) == 1 else None

def tokenize(text):
    sentences = sent_tokenize(text)
    return sentences

def sentiment_analyze(sentences, option):
    match option:
        case "VADER": 
            result = dict()
            sentpos = list()
            sentneg = list()
            sentneu = list()
            sentcom = list()
            analyzer = SentimentIntensityAnalyzer()
            for sentence in sentences:
                sentresult = analyzer.polarity_scores(sentence)
                sentpos.append(sentresult['pos'])
                sentneg.append(sentresult['neg'])
                sentneu.append(sentresult['neu'])
                sentcom.append(sentresult['compound'])
                
                avgpos = np.mean(sentpos)
                avgneg = np.mean(sentneg)
                avgneu = np.mean(sentneu)
                avgcom = np.mean(sentcom)
                result['pos'] = avgpos
                result['neg'] = avgneg
                result['neu'] = avgneu
                result['compound'] = avgcom
                return result
            
        case "TextBlob":
            result = dict()
            sentpol = list()
            sentsub = list()
            for sentence in sentences:
                blob = TextBlob(sentence)
                sentresult = blob.sentiment
                sentpol.append(sentresult[0])
                sentsub.append(sentresult[1])
                
                avgpol = np.mean(sentpol)
                avgsub = np.mean(sentsub)
                result['polarity'] = avgpol
                result['subjectivity'] = avgsub
                return result
        
        case "Flair":
            result = list()
            labelst = list()
            conflst = list()
            classifier = TextClassifier.load('en-sentiment')

            for senten in sentences:
                sentence = Sentence(senten)
                classifier.predict(sentence)
                for label in sentence.labels:
                        labelst.append(label.value)
                        conflst.append(label.score)
                    
                freqlabel = label_frequency(labelst)
                avgconf = np.mean(conflst)
                result.append(freqlabel)
                result.append(avgconf)
                return result
        
        case "SpaCy":
            result = dict()
            sentpol = list()
            sentsub = list()
            for sentence in sentences:
                doc = nlp(sentence)
                sentiment = doc._.blob.sentiment
                sentpol.append(sentiment.polarity)
                sentsub.append(sentiment.subjectivity)
                
                avgpol = np.mean(sentpol)
                avgsub = np.mean(sentsub)
                result['polarity'] = avgpol
                result['subjectivity'] = avgsub
                return result

def vader_analysis(text, option):
   datalst = list()
   data = list()
   result = sentiment_analyze(text, option)
   poscore = result['pos']
   negscore = result['neg']
   neuscore = result['neu']
   value = result['compound']
   print(result)
   if value > 0:
        category = 'POSITIVE'
   elif value < 0:
        category = 'NEGATIVE'
   else:
        category = 'NEUTRAL'

   data.append(category)
   data.append(poscore)
   data.append(negscore)
   data.append(neuscore)
   data.append(value)
   datalst.append(data)
   print(datalst)
   return datalst

def blob_analysis(text, option):
   datalst = list()
   data = list()
   result = sentiment_analyze(text, option)
   polscore = result['polarity']
   subscore = result['subjectivity']
   print(result)
   if polscore > 0:
       polabel = "Positive"
   elif polscore < 0:
       polabel = "Negative"
   else: 
       polabel = "Neutral"
   if subscore > 0.5:
       sublabel = "Subjective"
   elif subscore < 0.5:
       sublabel = "Objective"

   data.append(polabel)
   data.append(polscore)
   data.append(sublabel)
   data.append(subscore)
   datalst.append(data)
   print(datalst)
   return datalst

def flair_analysis(text, option):
   datalst = list()
   result = sentiment_analyze(text, option)
   datalst.append(result)
   print(datalst)
   return(datalst)

def spacy_analysis(text, option):
   datalst = list()
   data = list()
   result = sentiment_analyze(text, option)
   polscore = result['polarity']
   subscore = result['subjectivity']
   print(result)
   if polscore > 0:
       polabel = "Positive"
   elif polscore < 0:
       polabel = "Negative"
   else: 
       polabel = "Neutral"
   if subscore > 0.5:
       sublabel = "Subjective"
   elif subscore < 0.5:
       sublabel = "Objective"

   data.append(polabel)
   data.append(polscore)
   data.append(sublabel)
   data.append(subscore)
   datalst.append(data)
   print(datalst)

def write_excel(data, filename, sheetname, max_row, col):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    if sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
    else:
        sheet = workbook.create_sheet(sheetname)
    start_row = max_row + 1
    start_col = col
    for row_num, row_data in enumerate(data, start=start_row):
        for col_num, value in enumerate(row_data, start=start_col):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
    
    workbook.save(filename)
    print("Data written to file")

def write_file_excel(filename, sheetname, file, max_row):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    if sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
    else:
        sheet = workbook.create_sheet(sheetname)
    start_row = max_row + 1
    start_col = 1
    cell = sheet.cell(row = start_row, column = start_col)
    cell.value = file

    workbook.save(filename)

def process():
    if not folder_path:
        show_error(1)
    max_row = 1
    sheetname = input_sheet.get()
    fexcel = input_file.get()
    if not fexcel:
        show_error(2)
    col = get_colnum(fexcel, sheetname)
    for file in os.listdir(folder_path):
        fpath = os.path.join(folder_path, file)
        with open(fpath, "rb") as fh:
            fh = PDF.PdfReader(fh)
            num_pages = len(fh.pages)
            text = ""
            for i in range(num_pages):
                page = fh.pages[i]
                text += page.extract_text()

        writelst = list()
        sentences = tokenize(text)
        match selected_model:
            case "VADER":
                writelst = vader_analysis(sentences, selected_model)
            case "TextBLob":
                writelst = blob_analysis(sentences, selected_model)
            case "Flair":
                writelst = flair_analysis(sentences, selected_model)
            case "SpaCy":
                writelst = spacy_analysis(sentences, selected_model)
        write_file_excel(fexcel, sheetname, file, max_row)
        write_excel(writelst, fexcel, sheetname, max_row, col)
        max_row += 1
        print("done")
    output_label.config(text = "Data written to file!")

nlp = spacy.load("en_core_web_sm")
nlp.add_pipe("spacytextblob")

root = tk.Tk()
root.title("Sentiment Analysis!")
root.geometry("600x500")

title_label = tk.Label(root, text="Sentiment Analysis!", bg="#292828", fg="white", font=("Times New Roman", 20))
title_label.pack(side=tk.TOP, fill=tk.X, ipadx=10, ipady=20)

folder_button= tk.Button(root, text="Upload Folder", command=folder_directory, bg="#a8a5a5", fg="black", font=("Times New ROman", 14))
folder_button.pack(pady=5, ipadx=10, ipady=10)
file_label = tk.Label(root, text="Enter file name", bg="#a8a5a5", fg="black", font=("Times New Roman", 13))
file_label.pack(pady=10, ipadx=5, ipady=5, fill=tk.X)
input_file = tk.Entry(root, width=60)
input_file.pack(ipadx=5, ipady=5)
sheet_label = tk.Label(root, text="Enter sheet name", bg="#a8a5a5", fg="black", font=("Times New Roman", 12))
sheet_label.pack(pady=10, ipadx=5, ipady=5, fill=tk.X)
input_sheet = tk.Entry(root, width=60)
input_sheet.pack(ipadx=5, ipady=5)

menu = tk.Menu(root, tearoff=0)
menu.add_command(label="VADER", command=lambda: select_model("VADER"))
menu.add_command(label="TextBLob", command=lambda: select_model("TextBlob"))
menu.add_command(label="Flair", command=lambda: select_model("Flair"))
menu.add_command(label="SpaCy", command=lambda: select_model("SpaCy"))

menu_button = tk.Button(root, text="Model Select", command=lambda: menu.post(menu_button.winfo_rootx(), menu_button.winfo_rooty()), bg="#a8a5a5", fg="black", font=("Times New Roman", 12))
menu_button.pack(pady=10, ipadx=5, ipady=5)

output_label = tk.Label(root, text="")
output_label.pack()
button = tk.Button(root, text="Analyze Sentiments!!", command=process, bg="#5f5c5c", fg="white", font=("Times New Roman", 20))
button.pack(side=tk.BOTTOM, fill=tk.X, ipadx=10, ipady=10)

root.mainloop()